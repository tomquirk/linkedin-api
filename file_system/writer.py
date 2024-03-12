import os
import csv
import json
import logging
from openpyxl import load_workbook
import pandas as pd

DEFAULT_ENCODING = 'utf-8'
DEFAULT_VERBOSE = False
DEFAULT_EXCEL_SHEET_NAME = 'Sheet1'
DEFAULT_OVERWRITE = True


def get_field_names(data: list[dict], sort_keys: bool = False):
    key_dict = {}
    for row in data:
        for key in row.keys():
            key_dict[key] = True

    return sorted(list(key_dict.keys())) if sort_keys else list(key_dict.keys())


def write_json(file_path: str, data: dict, verbose=DEFAULT_VERBOSE, encoding=DEFAULT_ENCODING):
    with open(file_path, 'w', encoding=encoding) as jsonFile:
        json.dump(data, jsonFile, ensure_ascii=False, indent=2)

        if verbose:
            logging.info(f"Wrote JSON data to {file_path}")


def write_html(file_path, data, verbose=DEFAULT_VERBOSE, encoding=DEFAULT_ENCODING):
    with open(file_path, 'w', encoding=encoding) as file:
        file.writelines(data)

    if verbose:
        logging.info(
            f"Wrote HTML data to {file_path}")


def write_text(file_path, data, verbose=DEFAULT_VERBOSE, encoding=DEFAULT_ENCODING):
    with open(file_path, 'w', encoding=encoding) as file:
        file.writelines(data)

    if verbose:
        logging.info(
            f"Wrote TXT data to {file_path}")


def write_csv(file_path: str, data: list[dict], fieldnames=None, verbose=DEFAULT_VERBOSE, overwrite=DEFAULT_OVERWRITE, encoding=DEFAULT_ENCODING):
    if len(data) == 0:
        if verbose:
            logging.info("No CSV data to write")
        return

    write_mode = 'w' if overwrite or not os.path.isfile(file_path) else 'a'

    with open(file_path, write_mode, newline='', encoding=encoding) as csvfile:
        writer = csv.DictWriter(
            csvfile, fieldnames=fieldnames if fieldnames else get_field_names(data))

        if write_mode == 'w':
            writer.writeheader()

        for row in data:
            writer.writerow(row)

        if verbose:
            logging.info(
                f"{'Wrote' if write_mode == 'w' else 'Appended'} CSV data to {file_path} with {len(data)} rows")


def write_excel(file_path: str, data: list[dict], sheet_name=DEFAULT_EXCEL_SHEET_NAME, verbose=DEFAULT_VERBOSE, overwrite=DEFAULT_OVERWRITE, enable_str_conversion=False):
    """
    Append data to an existing Excel file or create a new sheet if it doesn't exist.

    Parameters:
    - file_path (str): Path to the existing Excel file.
    - sheet_name (str): Name of the sheet to append data to.
    - data (list of dicts or DataFrame): Data to be appended.
    """

    if len(data) == 0 and not overwrite:
        if verbose:
            logging.info("No EXCEL data to write")
        return

    df = pd.DataFrame(data)

    if enable_str_conversion:
        # Reference: https://stackoverflow.com/a/48454986
        df = df.applymap(lambda x: x.encode('unicode_escape').decode(
            'utf-8') if isinstance(x, str) else x)

    # If the file does not exist or need to be overwritten, create or overwrite it
    if not os.path.exists(file_path) or overwrite:
        df.to_excel(file_path, sheet_name=sheet_name, index=False)

        if verbose:
            logging.info(
                f"Wrote EXCEL data to {file_path} with {len(data)} rows")

        return

    # Load the workbook
    book = load_workbook(file_path)

    # Load and append the existing data
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = book

    # If the sheet name already exists, append the data
    if sheet_name in book.sheetnames:
        # Avoid overwriting existing content
        start_row = book[sheet_name].max_row
        # Avoid writing the header if the sheet isn't empty
        header = False if start_row > 0 else True
    # Otherwise, write the data from the first row with header
    else:
        start_row = 0
        header = True

    # Write the DataFrame to the Excel file
    df.to_excel(writer, sheet_name=sheet_name,
                startrow=start_row, header=header, index=False)

    # Save the changes
    writer.save()

    if verbose:
        logging.info(
            f"Wrote EXCEL data to {file_path} with {len(data)} rows")
